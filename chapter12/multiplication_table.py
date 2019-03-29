import sys
import openpyxl
from openpyxl.styles import Font


if len(sys.argv) > 1:

    WB = openpyxl.Workbook()
    WB.create_sheet(index=0, title="Test Sheet")
    SHEET = WB["Test Sheet"]
    FONT_OBJ = Font(bold=True)

    for row_num in range(2, int(sys.argv[1]) + 2):

        SHEET.cell(row=row_num, column=1).value = row_num - 1
        SHEET.cell(row=row_num, column=1).font = FONT_OBJ

    for column_num in range(2, int(sys.argv[1]) + 2):

        SHEET.cell(row=1, column=column_num).value = column_num - 1
        SHEET.cell(row=1, column=column_num).font = FONT_OBJ

    for row_num in range(2, int(sys.argv[1]) + 2):

        for column_num in range(2, int(sys.argv[1]) + 2):

            SHEET.cell(row=row_num, column=column_num).value = (row_num - 1) * (column_num - 1)

    WB.save("./chapter12/test.xlsx")

else:

    print("Number of CLI arguments is not correct.")
