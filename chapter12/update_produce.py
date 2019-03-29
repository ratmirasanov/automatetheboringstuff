#! python3
# update_produce.py -- Corrects costs in produce sales spreadsheet.

import openpyxl


WB = openpyxl.load_workbook("./chapter12/produce_sales.xlsx")
SHEET = WB["Sheet"]

# The produce types and their updated prices.
PRICE_UPDATES = {"Garlic": 3.07,
                 "Celery": 1.19,
                 "Lemon": 1.27}

# Loop through the rows and update the prices.

for row_num in range(2, SHEET.max_row):  # Skip the first row.

    produce_name = SHEET.cell(row=row_num, column=1).value

    if produce_name in PRICE_UPDATES:

        SHEET.cell(row=row_num, column=2).value = PRICE_UPDATES[produce_name]

WB.save("./chapter12/updated_produce_sales.xlsx")
